from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from sales_channels.exceptions import (
    MiraklImportInvalidFileLayoutError,
    MiraklImportInvalidFileTypeError,
)


@dataclass(slots=True)
class MiraklImportRow:
    filename: str
    row_number: int
    labels: list[str]
    codes: list[str]
    fields: dict[str, str]


@dataclass(slots=True)
class MiraklImportErrorRow:
    filename: str
    row_number: int
    fields: dict[str, str]


class MiraklWorkbookParser:
    DATA_SHEET_NAME = None
    ERROR_DETAILS_SHEET_NAME = "Error Details"

    def _open_workbook(self, *, export_file):
        filename = self._get_filename(export_file=export_file)
        extension = os.path.splitext(filename)[1].lower()
        if extension != ".xlsx":
            raise MiraklImportInvalidFileTypeError(
                f"Mirakl import export file '{filename or export_file.id}' must be an .xlsx file."
            )

        file_handle = export_file.file.open("rb")
        try:
            workbook = load_workbook(filename=file_handle, read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError, EOFError, KeyError, RuntimeError, TypeError) as exc:
            file_handle.close()
            raise MiraklImportInvalidFileLayoutError(
                f"Mirakl import export file '{filename or export_file.id}' could not be opened as a valid .xlsx workbook."
            ) from exc

        return filename, file_handle, workbook

    def get_layout(self, *, export_file) -> tuple[str, list[str], list[str]]:
        filename, file_handle, workbook = self._open_workbook(export_file=export_file)
        try:
            worksheet = workbook.active
            labels = self._normalize_row(row=next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None))
            codes = self._normalize_row(row=next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), None))
        finally:
            workbook.close()
            file_handle.close()

        if not codes:
            raise MiraklImportInvalidFileLayoutError(
                f"Mirakl import export file '{filename or export_file.id}' is missing the expected property-code row."
            )

        return filename, labels, codes

    def iter_rows(self, *, export_file) -> Iterator[MiraklImportRow]:
        filename, file_handle, workbook = self._open_workbook(export_file=export_file)
        try:
            worksheet = self._get_data_worksheet(workbook=workbook)
            labels = self._normalize_row(row=next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None))
            codes = self._normalize_row(row=next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), None))
            if not codes:
                raise MiraklImportInvalidFileLayoutError(
                    f"Mirakl import export file '{filename or export_file.id}' is missing the expected property-code row."
                )

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=3, values_only=True),
                start=3,
            ):
                fields = self._build_fields(codes=codes, row=row)
                if not fields:
                    continue

                yield MiraklImportRow(
                    filename=filename,
                    row_number=row_number,
                    labels=labels,
                    codes=codes,
                    fields=fields,
                )
        finally:
            workbook.close()
            file_handle.close()

    def iter_error_rows(self, *, export_file) -> Iterator[MiraklImportErrorRow]:
        filename, file_handle, workbook = self._open_workbook(export_file=export_file)
        try:
            worksheet = self._get_error_worksheet(workbook=workbook)
            if worksheet is None:
                return

            headers = self._normalize_error_headers(
                row=next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            )
            if not headers:
                return

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True),
                start=2,
            ):
                fields = self._build_fields(codes=headers, row=row)
                if not fields:
                    continue

                yield MiraklImportErrorRow(
                    filename=filename,
                    row_number=row_number,
                    fields=fields,
                )
        finally:
            workbook.close()
            file_handle.close()

    def count_rows(self, *, export_file) -> int:
        return sum(1 for _ in self.iter_rows(export_file=export_file))

    def _build_fields(self, *, codes: list[str], row) -> dict[str, str]:
        fields: dict[str, str] = {}
        if row is None:
            return fields

        has_value = False
        for index, raw_value in enumerate(row):
            if index >= len(codes):
                continue

            code = codes[index]
            if not code:
                continue

            value = self._normalize_value(value=raw_value)
            fields[code] = value
            if value != "":
                has_value = True

        if not has_value:
            return {}

        return fields

    def _normalize_row(self, *, row) -> list[str]:
        if row is None:
            return []
        return [self._normalize_value(value=value) for value in row]

    def _normalize_error_headers(self, *, row) -> list[str]:
        normalized = []
        for value in self._normalize_row(row=row):
            header = value.lower().replace("-", "_").replace(" ", "_")
            normalized.append(header)
        return normalized

    def _normalize_value(self, *, value) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _get_data_worksheet(self, *, workbook):
        if self.DATA_SHEET_NAME:
            return workbook[self.DATA_SHEET_NAME]
        return workbook.worksheets[0]

    def _get_error_worksheet(self, *, workbook):
        for worksheet in workbook.worksheets[1:]:
            if str(worksheet.title or "").strip() == self.ERROR_DETAILS_SHEET_NAME:
                return worksheet
        return None

    def _get_filename(self, *, export_file) -> str:
        return str(getattr(export_file.file, "name", "") or "")
