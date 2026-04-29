from __future__ import annotations

import csv
import io
import os
import re
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from sales_channels.integrations.mirakl.models import (
    MiraklEanCode,
    MiraklProduct,
    MiraklProductIssue,
    MiraklProperty,
    MiraklSalesChannelFeedItem,
)
from sales_channels.integrations.mirakl.utils.offer_fields import add_offer_field_aliases


class MiraklTransformationErrorReportIssueSyncFactory:
    """Parse Mirakl product import issue reports and persist product issues."""

    SOURCE_ERROR_REPORT_ERROR = "error_report_error"
    SOURCE_ERROR_REPORT_WARNING = "error_report_warning"
    SOURCE_TRANSFORMATION_ERROR = "transformation_error_report_error"
    SOURCE_TRANSFORMATION_WARNING = "transformation_error_report_warning"

    def __init__(self, *, feed) -> None:
        self.feed = feed
        self.sales_channel = self._resolve_sales_channel_instance(feed=feed)
        self.sku_headers = self._get_candidate_sku_headers()
        self.ean_headers = self._get_candidate_ean_headers()
        self._cleared_issue_source_keys: set[tuple[int, str]] = set()

    def run(self) -> int:
        remote_product_lookup = self._build_remote_product_lookup()
        touched_remote_product_ids: set[int] = set()
        synced_issues = self._sync_report(
            file_field=self.feed.error_report_file,
            remote_product_lookup=remote_product_lookup,
            touched_remote_product_ids=touched_remote_product_ids,
            source_error=self.SOURCE_ERROR_REPORT_ERROR,
            source_warning=self.SOURCE_ERROR_REPORT_WARNING,
        )
        synced_issues += self._sync_report(
            file_field=self.feed.transformation_error_report_file,
            remote_product_lookup=remote_product_lookup,
            touched_remote_product_ids=touched_remote_product_ids,
            source_error=self.SOURCE_TRANSFORMATION_ERROR,
            source_warning=self.SOURCE_TRANSFORMATION_WARNING,
        )

        self._refresh_remote_product_statuses(
            remote_product_ids=sorted(touched_remote_product_ids),
        )
        return synced_issues

    def _sync_report(
        self,
        *,
        file_field,
        remote_product_lookup: dict[str, MiraklProduct],
        touched_remote_product_ids: set[int],
        source_error: str,
        source_warning: str,
    ) -> int:
        if not file_field:
            return 0

        try:
            with file_field.open("rb") as file_handle:
                content = file_handle.read()
        except OSError:
            return 0
        if not content:
            return 0

        synced_issues = 0
        for row_index, row in self._iter_report_rows(file_field=file_field, content=content):
            if not row:
                continue
            remote_product = self._resolve_remote_product(
                row=row,
                remote_product_lookup=remote_product_lookup,
            )
            if remote_product is None:
                continue

            synced_issues += self._upsert_issues(
                remote_product=remote_product,
                row=row,
                row_index=row_index,
                source_error=source_error,
                source_warning=source_warning,
            )
            touched_remote_product_ids.add(remote_product.id)
        return synced_issues

    def _iter_report_rows(
        self,
        *,
        file_field,
        content: bytes,
    ) -> Iterable[tuple[int, dict[str, str]]]:
        filename = str(getattr(file_field, "name", "") or "")
        extension = os.path.splitext(filename)[1].lower()
        if extension in {".xlsx", ".xlsm"} or content.startswith(b"PK"):
            rows = list(self._iter_workbook_rows(content=content))
            if rows:
                return rows
            return list(self._iter_csv_rows(content=content))

        rows = list(self._iter_csv_rows(content=content))
        if rows:
            return rows
        return list(self._iter_workbook_rows(content=content))

    def _iter_csv_rows(self, *, content: bytes) -> Iterable[tuple[int, dict[str, str]]]:
        text = self._decode_csv_bytes(content=content)
        if not text.strip():
            return []

        rows: list[tuple[int, dict[str, str]]] = []
        dialect = self._detect_csv_dialect(text=text)
        for row_index, raw_row in enumerate(csv.DictReader(io.StringIO(text), dialect=dialect), start=1):
            row = self._normalize_mapping_row(raw_row=raw_row)
            if row:
                rows.append((row_index, row))
        return rows

    def _iter_workbook_rows(self, *, content: bytes) -> Iterable[tuple[int, dict[str, str]]]:
        try:
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError, EOFError, KeyError, RuntimeError, TypeError):
            return []

        rows: list[tuple[int, dict[str, str]]] = []
        try:
            worksheet = workbook.active
            row_iter = worksheet.iter_rows(values_only=True)
            headers = next(row_iter, None)
            if not headers:
                return []

            normalized_headers = [self._normalize_header(value=header) for header in headers]
            for row_index, values in enumerate(row_iter, start=1):
                row = self._build_row(headers=headers, normalized_headers=normalized_headers, values=values)
                if row:
                    rows.append((row_index, row))
        finally:
            workbook.close()
        return rows

    def _resolve_sales_channel_instance(self, *, feed):
        sales_channel = feed.sales_channel
        get_real_instance = getattr(sales_channel, "get_real_instance", None)
        if callable(get_real_instance):
            return get_real_instance()
        return sales_channel

    def _build_remote_product_lookup(self) -> dict[str, MiraklProduct]:
        lookup: dict[str, MiraklProduct] = {}
        queryset = MiraklSalesChannelFeedItem.objects.filter(feed=self.feed).select_related("remote_product")
        for item in queryset.iterator():
            remote_product = self._resolve_mirakl_product(remote_product=item.remote_product)
            if remote_product is None:
                continue

            lookup_key_candidates = [str(remote_product.remote_sku or "").strip()]
            for row in item.payload_data or []:
                if not isinstance(row, dict):
                    continue
                normalized_row = add_offer_field_aliases(row={
                    self._normalize_header(value=key): self._stringify(value=value)
                    for key, value in row.items()
                })
                lookup_key_candidates.extend(normalized_row.get(header, "") for header in self.sku_headers)
                lookup_key_candidates.extend(normalized_row.get(header, "") for header in self.ean_headers)

            for candidate in lookup_key_candidates:
                normalized_candidate = self._normalize_lookup_value(value=candidate)
                if normalized_candidate:
                    lookup[normalized_candidate] = remote_product

        return lookup

    def _get_candidate_sku_headers(self) -> list[str]:
        return self._get_candidate_headers(
            representation_types=(
                MiraklProperty.REPRESENTATION_PRODUCT_SKU,
                MiraklProperty.REPRESENTATION_PRODUCT_CONFIGURABLE_SKU,
            ),
        )

    def _get_candidate_ean_headers(self) -> list[str]:
        return self._get_candidate_headers(
            representation_types=(MiraklProperty.REPRESENTATION_PRODUCT_EAN,),
        )

    def _get_candidate_headers(self, *, representation_types: tuple[str, ...]) -> list[str]:
        headers: list[str] = []
        if self.feed.product_type_id:
            headers.extend(
                self.feed.product_type.items.filter(
                    remote_property__representation_type__in=representation_types,
                )
                .select_related("remote_property")
                .values_list("remote_property__code", flat=True)
            )

        headers.extend(
            MiraklProperty.objects.filter(
                sales_channel=self.sales_channel,
                representation_type__in=representation_types,
            ).values_list("code", flat=True)
        )

        return list(
            dict.fromkeys(
                self._normalize_header(value=header)
                for header in headers
                if self._normalize_header(value=header)
            )
        )

    def _build_row(self, *, headers, normalized_headers: list[str], values) -> dict[str, str]:
        row: dict[str, str] = {}
        if values is None:
            return row
        for index, normalized_header in enumerate(normalized_headers):
            if not normalized_header:
                continue
            raw_value = values[index] if index < len(values) else ""
            row[normalized_header] = self._stringify(value=raw_value)
        return row

    def _resolve_remote_product(
        self,
        *,
        row: dict[str, str],
        remote_product_lookup: dict[str, MiraklProduct],
    ) -> MiraklProduct | None:
        for header in self.sku_headers:
            candidate = self._normalize_lookup_value(value=row.get(header))
            if not candidate:
                continue
            matched = remote_product_lookup.get(candidate)
            if matched is not None:
                return matched
            matched = self._match_remote_product_by_sku(identifier=candidate)
            if matched is not None:
                return matched

        for header in self.ean_headers:
            candidate = self._normalize_lookup_value(value=row.get(header))
            if not candidate:
                continue
            matched = remote_product_lookup.get(candidate)
            if matched is not None:
                return matched
            matched = self._match_remote_product_by_ean(identifier=candidate)
            if matched is not None:
                return matched

        return None

    def _match_remote_product_by_sku(self, *, identifier: str) -> MiraklProduct | None:
        return (
            MiraklProduct.objects.filter(
                sales_channel=self.sales_channel,
                remote_sku=identifier,
            )
            .order_by("id")
            .first()
        )

    def _match_remote_product_by_ean(self, *, identifier: str) -> MiraklProduct | None:
        ean_match = (
            MiraklEanCode.objects.filter(
                remote_product__sales_channel=self.sales_channel,
                ean_code=identifier,
            )
            .select_related("remote_product")
            .order_by("id")
            .first()
        )
        if ean_match is None or ean_match.remote_product is None:
            return None
        return self._resolve_mirakl_product(remote_product=ean_match.remote_product)

    def _resolve_mirakl_product(self, *, remote_product) -> MiraklProduct | None:
        resolved = remote_product
        get_real_instance = getattr(resolved, "get_real_instance", None)
        if callable(get_real_instance):
            resolved = get_real_instance()
        return resolved if isinstance(resolved, MiraklProduct) else None

    def _upsert_issues(
        self,
        *,
        remote_product: MiraklProduct,
        row: dict[str, str],
        row_index: int,
        source_error: str,
        source_warning: str,
    ) -> int:
        self._clear_existing_report_issues(
            remote_product=remote_product,
            sources=(source_error, source_warning),
        )
        created_or_updated = 0
        line_number = row.get("line_number") or row.get("line_number_") or row.get("line")
        for source, severity, column_name, is_rejected in (
            (source_error, "ERROR", "errors", source_error == self.SOURCE_ERROR_REPORT_ERROR),
            (source_warning, "WARNING", "warnings", False),
        ):
            for entry_index, (code, message) in enumerate(self._parse_issue_entries(value=row.get(column_name, "")), start=1):
                self._create_issue(
                    remote_product=remote_product,
                    source=source,
                    severity=severity,
                    is_rejected=is_rejected,
                    code=code,
                    message=message,
                    line_number=line_number,
                    row_index=row_index,
                    column_name=column_name,
                    entry_index=entry_index,
                    row=row,
                )
                created_or_updated += 1
        return created_or_updated

    def _clear_existing_report_issues(
        self,
        *,
        remote_product: MiraklProduct,
        sources: tuple[str, str],
    ) -> None:
        for source in sources:
            key = (remote_product.id, source)
            if key in self._cleared_issue_source_keys:
                continue
            MiraklProductIssue.objects.filter(
                remote_product=remote_product,
                raw_data__source=source,
            ).delete()
            self._cleared_issue_source_keys.add(key)

    def _create_issue(
        self,
        *,
        remote_product: MiraklProduct,
        source: str,
        severity: str,
        is_rejected: bool,
        code: str,
        message: str | None,
        line_number: str | None,
        row_index: int,
        column_name: str,
        entry_index: int,
        row: dict[str, str],
    ) -> None:
        raw_data = {
            "source": source,
            "feed_id": self.feed.id,
            "line_number": line_number,
            "row_index": row_index,
            "column_name": column_name,
            "entry_index": entry_index,
            "row": row,
        }
        issue = MiraklProductIssue.objects.create(
            multi_tenant_company=remote_product.multi_tenant_company,
            remote_product=remote_product,
            main_code=code,
            code=code,
            message=message,
            severity=severity,
            reason_label=None,
            attribute_code=None,
            is_rejected=is_rejected,
            raw_data=raw_data,
        )

        if self.feed.sales_channel_view_id:
            issue.views.set([self.feed.sales_channel_view])
        else:
            issue.views.clear()

    def _parse_issue_entries(self, *, value: str) -> Iterable[tuple[str, str | None]]:
        text = self._stringify(value=value)
        if not text:
            return []

        entries: list[tuple[str, str | None]] = []
        for chunk in [part.strip() for part in text.split(",") if str(part or "").strip()]:
            if "|" in chunk:
                code, message = chunk.split("|", 1)
            else:
                code, message = chunk, ""
            normalized_code = str(code or "").strip()
            if not normalized_code:
                continue
            normalized_message = str(message or "").strip() or None
            entries.append((normalized_code, normalized_message))
        return entries

    def _normalize_header(self, *, value) -> str:
        text = self._stringify(value=value).lower()
        text = text.replace(" ", "_")
        text = re.sub(r"[^a-z0-9_]+", "_", text)
        return text.strip("_")

    def _normalize_mapping_row(self, *, raw_row) -> dict[str, str]:
        if not isinstance(raw_row, dict):
            return {}
        ordered_headers = [header for header in raw_row.keys() if header is not None]
        row = {
            self._normalize_header(value=header): self._stringify(value=raw_row.get(header))
            for header in ordered_headers
        }
        self._recover_overflow_tail_values(
            row=row,
            ordered_headers=ordered_headers,
            overflow_values=raw_row.get(None),
        )
        if any(value for value in row.values()):
            return row
        return {}

    def _recover_overflow_tail_values(
        self,
        *,
        row: dict[str, str],
        ordered_headers: list[str],
        overflow_values,
    ) -> None:
        if not ordered_headers or not isinstance(overflow_values, list) or not overflow_values:
            return

        normalized_headers = [
            self._normalize_header(value=header)
            for header in ordered_headers
        ]
        tail_values = [self._stringify(value=row.get(normalized_headers[-1]))]
        tail_values.extend(self._stringify(value=value) for value in overflow_values)
        recover_count = min(len(normalized_headers), len(tail_values))
        if recover_count <= 0:
            return

        for header, value in zip(normalized_headers[-recover_count:], tail_values[-recover_count:]):
            if not header:
                continue
            row[header] = value

    def _normalize_lookup_value(self, *, value) -> str:
        return self._stringify(value=value).strip()

    def _stringify(self, *, value) -> str:
        if value in (None, ""):
            return ""
        return str(value)

    def _decode_csv_bytes(self, *, content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return content.decode("utf-8", errors="ignore")

    def _detect_csv_dialect(self, *, text: str):
        sample = text[:4096]
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            return csv.excel

    def _refresh_remote_product_statuses(self, *, remote_product_ids: list[int]) -> None:
        if not remote_product_ids:
            return

        for remote_product in MiraklProduct.objects.filter(id__in=remote_product_ids).order_by("id").iterator():
            remote_product.refresh_status(commit=True)
