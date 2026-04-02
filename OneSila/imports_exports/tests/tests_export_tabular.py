import csv
from io import BytesIO, StringIO
from types import SimpleNamespace

from django.test import SimpleTestCase
from openpyxl import load_workbook

from imports_exports.factories.exports.tabular import (
    build_csv_export_content,
    build_excel_export_content,
    build_tabular_export,
)


class TabularExportBuilderTest(SimpleTestCase):
    def test_build_tabular_export_flattens_nested_product_data(self):
        export_process = SimpleNamespace(
            kind="products",
            raw_data=[
                {
                    "sku": "PARENT-1",
                    "name": "Parent Product",
                    "translations": [
                        {
                            "language": "en",
                            "name": "Parent Product",
                            "sales_channel": None,
                            "bullet_points": ["One", "Two"],
                        },
                        {
                            "language": "en",
                            "name": "Channel Product",
                            "sales_channel": 12,
                        },
                    ],
                    "properties": [
                        {
                            "property": {
                                "internal_name": "color",
                            },
                            "value": "Red",
                            "requirement": "OPTIONAL",
                        }
                    ],
                    "images": [
                        {
                            "image_url": "https://example.com/image.jpg",
                            "type": "PACK",
                        }
                    ],
                    "variations": [
                        {
                            "variation_data": {
                                "sku": "CHILD-1",
                                "type": "SIMPLE",
                                "translations": [
                                    {
                                        "language": "en",
                                        "name": "Child Product",
                                        "sales_channel": None,
                                    }
                                ],
                            }
                        }
                    ],
                }
            ],
        )

        headers, rows = build_tabular_export(export_process=export_process)

        self.assertIn("translation_name_en_sc_default", headers)
        self.assertIn("translation_name_en_sc_12", headers)
        self.assertIn("translation_bullet_points_en_sc_default", headers)
        self.assertIn("color", headers)
        self.assertIn("color_requirement", headers)
        self.assertIn("image_1_url", headers)
        self.assertIn("configurable_parent_sku", headers)
        self.assertIn("configurable_product_sku", headers)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["sku"], "PARENT-1")
        self.assertEqual(rows[0]["color"], "Red")
        self.assertEqual(rows[1]["sku"], "CHILD-1")
        self.assertEqual(rows[1]["configurable_parent_sku"], "PARENT-1")

    def test_build_csv_export_content_writes_headers_and_rows(self):
        export_process = SimpleNamespace(
            kind="ean_codes",
            raw_data=[
                {
                    "ean_code": "1234567890123",
                    "product_sku": "SKU-1",
                }
            ],
        )

        payload = build_csv_export_content(export_process=export_process).decode("utf-8-sig")
        rows = list(csv.reader(StringIO(payload)))

        self.assertEqual(rows[0], ["ean_code", "product_sku"])
        self.assertEqual(rows[1], ["1234567890123", "SKU-1"])

    def test_build_excel_export_content_writes_headers_and_rows(self):
        export_process = SimpleNamespace(
            kind="ean_codes",
            raw_data=[
                {
                    "ean_code": "1234567890123",
                    "product_sku": "SKU-1",
                }
            ],
        )

        payload = build_excel_export_content(export_process=export_process)
        workbook = load_workbook(filename=BytesIO(payload))
        worksheet = workbook.active

        self.assertEqual(
            [worksheet.cell(row=1, column=1).value, worksheet.cell(row=1, column=2).value],
            ["ean_code", "product_sku"],
        )
        self.assertEqual(
            [worksheet.cell(row=2, column=1).value, worksheet.cell(row=2, column=2).value],
            ["1234567890123", "SKU-1"],
        )
